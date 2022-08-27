from django.shortcuts import render,redirect,HttpResponse
from openpyxl import load_workbook

from app01 import models
from app01.utils.Pagination import Pagination

def grade_list(request):
    """成绩罗列"""
    info = request.session.get('info')
    teacher_id = request.session['info']['id']
    if not info:
        redirect('/login/')

    #1. 数据库中获取所有部门列表
    #[对象，对象，对象]
    queryset = models.Grade.objects.filter(teacher_id=teacher_id)
    page_obj = Pagination(request, queryset, page_size=5)
    context = {
        'queryset': page_obj.page_queryset,
        "page_string": page_obj.html(),
    }

    return render(request, 'grade_list.html', context)

def grade_multi(request):
    """批量上传文件"""
    # 1.获取用户上传的文件对象
    file_obj = request.FILES.get('exc')
    # print(type(file_obj))

    # 2. 对象传递给openpyxl，由它打开读取内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    # 3. 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        year = row[0].value
        term = row[1].value
        course_id = row[2].value
        teacher_id = request.session['info']['id']
        student_id = row[3].value
        grade = row[4].value

        print(year)
        print(student_id)
        print(grade)
        exists = models.Grade.objects.filter(course_id=course_id, student_id=student_id).exists()
        if not exists:
            models.Grade.objects.create(grade=grade, year=year, term=term, course_id=course_id, student_id=student_id, teacher_id=teacher_id)

    return redirect('/grade/list/')

def grade_search(request):
    """学生成绩查询"""
    year = request.POST.get('year')
    term = request.POST.get('term')
    student_id = request.session['info']['id']

    search_dict = dict()
    if year:
        search_dict['year'] = year
    if term:
        search_dict['term'] = term
    if student_id:
        search_dict['student_id'] = student_id

    queryset = models.Grade.objects.filter(**search_dict)
    print(queryset)
    page_obj = Pagination(request, queryset, page_size=5)
    context = {
        'queryset': page_obj.page_queryset,
        "page_string": page_obj.html(),
    }

    return render(request, 'grade_search.html', context)
